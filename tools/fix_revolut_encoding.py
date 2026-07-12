#!/usr/bin/env python3
"""Fix the double-encoded (mojibake) text in a Revolut XLSX statement export.

Revolut's statement export declares UTF-8 but writes text as double-encoded
UTF-8: the original UTF-8 bytes are reinterpreted as Latin-1/CP1252 and
re-encoded to UTF-8. This recovers the original text (e.g. ``zahÃ¡jenÃ­`` ->
``zahájení``) and writes a clean CSV (or XLSX).
"""

import argparse
import csv
from pathlib import Path

import openpyxl

from backend.routers.utils.revolut import fix_mojibake


def read_fixed_rows(source):
    """Yield rows from the XLSX with every string cell de-mojibaked."""
    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    worksheet = workbook.active
    for row in worksheet.iter_rows(values_only=True):
        yield [fix_mojibake(cell) for cell in row]
    workbook.close()


def write_csv(rows, destination):
    with destination.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerows(rows)


def write_xlsx(rows, destination):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    for row in rows:
        worksheet.append(row)
    workbook.save(destination)


parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
parser.add_argument("source", type=Path, help="Path to the Revolut .xlsx export")
parser.add_argument(
    "-o",
    "--output",
    type=Path,
    help="Output path (defaults to '<source>.fixed.csv')",
)
parser.add_argument(
    "-f",
    "--format",
    choices=("csv", "xlsx"),
    help="Output format (defaults to the output extension, else csv)",
)
args = parser.parse_args()

if not args.source.is_file():
    parser.error(f"file '{args.source}' not found")

output = args.output
fmt = args.format
if output is None:
    fmt = fmt or "csv"
    output = args.source.with_suffix(f".fixed.{fmt}")
elif fmt is None:
    fmt = "xlsx" if output.suffix.lower() == ".xlsx" else "csv"

rows = list(read_fixed_rows(args.source))

if fmt == "xlsx":
    write_xlsx(rows, output)
else:
    write_csv(rows, output)

print(f"Wrote {len(rows)} rows to {output}")
